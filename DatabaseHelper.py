import zipfile
import os
from PyPDF2 import PdfFileReader
from openpyxl import Workbook, load_workbook


NoneType = type(None)
UPLOAD_FOLDER = r'Uploaded Documnet'

def fileUploded(user_sub):
    """
        parameters:- user_sub
        returns all documents in folder
    """
    files = []
    for filename in os.listdir(user_sub):
        files.append(filename)
    return files


def pageCounter(file):
    """
        parameters:- file - path of file
        count number of pages in file
    """
    temp = open(file, "rb")
    pdf = PdfFileReader(temp)
    pages = pdf.getNumPages()
    temp.close()
    print("document ", file, "pages ", pages)

    return pages

def createBillExcel(bill):
    """
        parameters:- bill - path of bill xcel files
        create empty document
    """
    wb = Workbook()
    wb.save(bill)
    wb.close()
    print("--------Bill file created--------")
    return None

def updatedBill(bill, users, record):
    """
        parameters:- bill - path of bill xcel files, users :- sorted list of user send doc to print,
        record dict has record of users and how many pages they want to print
    """
    wb = load_workbook(bill)
    sheet = wb["Sheet"]
    sheet.cell(row=1, column=1).value = "UserName"
    sheet.cell(row=1, column=2).value = "Pages"
    sheet.cell(row=1, column=3).value = "Bill"
    r = 2
    for entry in users:
        sheet.cell(row=r, column=1).value = entry
        sheet.cell(row=r, column=2).value = record[entry]
        sheet.cell(row=r, column=3).value = record[entry] * 1
        r += 1
    wb.save(bill)
    wb.close()
    print("--------Bill file Updated--------")
    return None

def createZip(user_sub, files):
    """
        parameter:- folder path , files list
        This will create zip file
    """
    zipf = zipfile.ZipFile('Document.zip', 'w', zipfile.ZIP_DEFLATED)
    for file in files:
        zipf.write(user_sub + file)
    zipf.close()
    print("--------zip file created --------")
    return None


def deleteFiles(user_sub):
    """
        parameter:- folder path
        This will delete files from folder
    """
    for filename in os.listdir(user_sub):
        os.remove(user_sub + filename)
        print("deleted ", filename)
    return None
